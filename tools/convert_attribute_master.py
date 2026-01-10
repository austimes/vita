#!/usr/bin/env python3
"""
Convert veda-attribute-master-export.xlsx to attribute-master.json.

This is a one-off script to convert the VEDA attribute master export
into a JSON format suitable for runtime validation in table_schemas.py.

Usage:
    uv run python tools/convert_attribute_master.py

The output file is written to vedalang/schema/attribute-master.json.
"""

import json
import re
from pathlib import Path

import openpyxl


def parse_indexes(indexes_str: str) -> list[str]:
    """Parse indexes string like '(r,datayear,p,c)' into ['r', 'datayear', 'p', 'c']."""
    if not indexes_str:
        return []
    # Remove parentheses and split by comma
    match = re.match(r"\(([^)]*)\)", indexes_str.strip())
    if match:
        inner = match.group(1)
        return [idx.strip() for idx in inner.split(",") if idx.strip()]
    return []


def convert_attribute_master():
    """Convert Excel attribute master to JSON format."""
    input_path = Path("veda/veda-attribute-master-export.xlsx")
    output_path = Path("vedalang/schema/attribute-master.json")

    wb = openpyxl.load_workbook(input_path, read_only=True)
    sheet = wb.active

    # Get header row
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Find column indices
    col_idx = {h: i for i, h in enumerate(headers) if h}

    attributes = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        attr_name = row[col_idx["Attribute"]]
        if not attr_name:
            continue

        # Parse aliases (comma-separated) - need this first for column_headers
        alias_str = row[col_idx.get("Alias", -1)] or ""
        aliases = []
        if alias_str:
            aliases = [a.strip() for a in alias_str.split(",") if a.strip()]

        # Build list of valid column headers (canonical + all aliases, lowercase)
        column_headers = [attr_name.lower()]
        column_headers.extend(a.lower() for a in aliases)

        # Parse indexes into structured list
        indexes_raw = row[col_idx.get("Indexes", -1)] or ""
        indexes = parse_indexes(indexes_raw)

        # Build attribute entry with ALL columns
        entry = {
            "column_header": attr_name.lower(),  # Primary/canonical header
            "column_headers": column_headers,  # All valid headers (including aliases)
            "description": row[col_idx.get("Description", -1)] or "",
            "time_series": row[col_idx.get("TimeSeries", -1)] == "Yes",
            "process": row[col_idx.get("Process", -1)] == "T",
            "commodity": row[col_idx.get("Commodity", -1)] == "T",
            "timeslice": row[col_idx.get("TimeSlice", -1)] or "",
            "limtype": row[col_idx.get("LimType", -1)] or "",
            "currency": row[col_idx.get("Currency", -1)] == "CUR",
            "stage": row[col_idx.get("Stage", -1)] == "T",
            "sow": row[col_idx.get("Sow", -1)] == "T",
        }

        if aliases:
            entry["aliases"] = aliases

        # Add indexes (structured and raw)
        if indexes_raw:
            entry["indexes"] = indexes
            entry["indexes_raw"] = indexes_raw

        # Add OtherIndexes if present
        other_indexes = row[col_idx.get("OtherIndexes", -1)] or ""
        if other_indexes:
            entry["other_indexes"] = other_indexes

        # Add remaining columns if they have values
        related = row[col_idx.get("RelatedSetsAndParameters", -1)] or ""
        if related:
            entry["related_sets_and_parameters"] = related

        units_ranges = row[col_idx.get("UnitsRangesAndDefaultValuesAndDefaultInterExtrapolation", -1)] or ""
        if units_ranges:
            entry["units_ranges_defaults"] = units_ranges

        instances = row[col_idx.get("InstancesRequiredOmitSpecialConditions", -1)] or ""
        if instances:
            entry["instances_conditions"] = instances

        affected = row[col_idx.get("AffectedEquationsOrVariables", -1)] or ""
        if affected:
            entry["affected_equations_or_variables"] = affected

        attributes[attr_name] = entry

    # Write JSON output
    output = {
        "_comment": "Auto-generated from veda-attribute-master-export.xlsx. Do not edit.",
        "_source": "tools/convert_attribute_master.py",
        "attributes": attributes,
    }

    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)

    print(f"Converted {len(attributes)} attributes to {output_path}")
    return attributes


if __name__ == "__main__":
    convert_attribute_master()
