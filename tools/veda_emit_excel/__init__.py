"""TableIR to Excel emitter for VEDA tables."""

import json
from pathlib import Path

import jsonschema
import yaml
from openpyxl import Workbook

from vedalang.compiler.online_compat import validate_online_compat

SCHEMA_PATH = (
    Path(__file__).parent.parent.parent / "vedalang" / "schema" / "tableir.schema.json"
)

# Scalar tags that should NOT have a header row - values are emitted directly
SCALAR_TAGS = {"~STARTYEAR", "~ACTIVEPDEF"}


def load_schema() -> dict:
    """Load the TableIR JSON schema."""
    with open(SCHEMA_PATH) as f:
        return json.load(f)


def validate_tableir(tableir: dict) -> None:
    """Validate TableIR against schema. Raises jsonschema.ValidationError if invalid."""
    schema = load_schema()
    jsonschema.validate(tableir, schema)


def emit_excel(tableir: dict, out_dir: Path, validate: bool = True) -> list[Path]:
    """
    Convert TableIR dict to Excel files.

    Args:
        tableir: TableIR dictionary with files/sheets/tables structure
        out_dir: Directory to write Excel files to
        validate: Whether to validate against schema first

    Returns:
        List of paths to created Excel files
    """
    if validate:
        validate_tableir(tableir)

        online_errors = validate_online_compat(tableir)
        if online_errors:
            raise ValueError(
                "VEDA Online compatibility errors:\n"
                + "\n".join(f"  - {e}" for e in online_errors)
            )

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    created_files = []

    for file_spec in tableir.get("files", []):
        file_path = out_dir / file_spec["path"]
        file_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_spec in file_spec.get("sheets", []):
            ws = wb.create_sheet(title=sheet_spec["name"])

            current_row = 1
            for note in sheet_spec.get("notes", []):
                ws.cell(row=current_row, column=1, value=note)
                current_row += 1
            if sheet_spec.get("notes"):
                current_row += 1

            for table in sheet_spec.get("tables", []):
                # Emit ~UC_SETS declarations before the table tag if present
                uc_sets = table.get("uc_sets", {})
                for uc_key, uc_value in uc_sets.items():
                    # Format: ~UC_SETS: R_E: AllRegions or ~UC_SETS: T_E
                    # Note: Empty values should not have trailing space
                    if uc_value:
                        uc_sets_cell = f"~UC_SETS: {uc_key}: {uc_value}"
                    else:
                        uc_sets_cell = f"~UC_SETS: {uc_key}"
                    ws.cell(row=current_row, column=1, value=uc_sets_cell)
                    current_row += 1

                tag = table["tag"]
                ws.cell(row=current_row, column=1, value=tag)
                current_row += 1

                rows = table.get("rows", [])
                if rows:
                    # Check if this is a scalar tag (no header row needed)
                    is_scalar = tag in SCALAR_TAGS

                    if is_scalar:
                        # Scalar tags: emit values directly without header row
                        # Rows should have single "value" key
                        for row in rows:
                            extra_keys = set(row.keys()) - {"value"}
                            if extra_keys:
                                raise ValueError(
                                    f"Scalar tag {tag} rows must only have "
                                    f"'value' key, found: {extra_keys}"
                                )
                            value = row.get("value")
                            if value is not None:
                                ws.cell(row=current_row, column=1, value=value)
                            current_row += 1
                    else:
                        # Normal tables: collect columns and emit header + data
                        columns = []
                        for row in rows:
                            for key in row.keys():
                                if key not in columns:
                                    columns.append(key)

                        for col_idx, col_name in enumerate(columns, start=1):
                            ws.cell(row=current_row, column=col_idx, value=col_name)
                        current_row += 1

                        for row in rows:
                            for col_idx, col_name in enumerate(columns, start=1):
                                val = row.get(col_name)
                                if val is not None:
                                    ws.cell(row=current_row, column=col_idx, value=val)
                            current_row += 1

                current_row += 1

        wb.save(file_path)
        created_files.append(file_path)

    return created_files


def load_tableir(path: Path) -> dict:
    """Load TableIR from YAML or JSON file."""
    path = Path(path)
    with open(path) as f:
        if path.suffix in (".yaml", ".yml"):
            return yaml.safe_load(f)
        else:
            return json.load(f)
