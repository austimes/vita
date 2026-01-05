#!/usr/bin/env python3
"""
Compare other_indexes definitions between xl2times/config/times-info.json
and veda/Attribute-master-export.xlsx.

This tool identifies discrepancies in which indices should go into the
other_indexes column for each TIMES attribute.

Usage:
    uv run python tools/compare_other_indexes.py
"""

import json
from pathlib import Path

import openpyxl


def load_times_info() -> dict[str, dict]:
    """Load times-info.json and extract other_indexes mappings."""
    path = Path("xl2times/config/times-info.json")
    with open(path) as f:
        data = json.load(f)

    result = {}
    for entry in data:
        name = entry["name"]
        indexes = entry.get("indexes", [])
        mapping = entry.get("mapping", [])

        # Find indices that map to "other_indexes"
        other_idx = []
        for i, m in enumerate(mapping):
            if m == "other_indexes" and i < len(indexes):
                other_idx.append(indexes[i].lower())

        result[name] = {
            "indexes": [idx.lower() for idx in indexes],
            "mapping": mapping,
            "other_indexes": other_idx,
        }

    return result


def load_attribute_master() -> dict[str, dict]:
    """Load Attribute-master-export.xlsx and extract OtherIndexes."""
    path = Path("veda/Attribute-master-export.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = wb.active

    # Get headers
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    result = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        attr_name = row[col_idx["Attribute"]]
        if not attr_name:
            continue

        other_indexes_str = row[col_idx.get("OtherIndexes", -1)] or ""
        indexes_str = row[col_idx.get("Indexes", -1)] or ""

        # Parse other_indexes (could be comma-separated)
        other_idx = []
        if other_indexes_str:
            other_idx = [x.strip().lower() for x in other_indexes_str.split(",") if x.strip()]

        result[attr_name] = {
            "indexes": indexes_str,
            "other_indexes": other_idx,
        }

    return result


def normalize_index_name(name: str) -> str:
    """Normalize index names for comparison."""
    # Common mappings between times-info naming and attribute-master naming
    mappings = {
        "upt": "commodity_group",
        "cg": "commodity_group",
        "ie": "ie",
        "io": "io",
        "item": "item",
        "reg2": "region2",
        "com2": "commodity2",
        "ts2": "timeslice2",
    }
    return mappings.get(name.lower(), name.lower())


def compare_other_indexes():
    """Compare other_indexes between the two sources."""
    times_info = load_times_info()
    attr_master = load_attribute_master()

    print("=" * 80)
    print("Comparing other_indexes between times-info.json and Attribute-master-export.xlsx")
    print("=" * 80)
    print()

    discrepancies = []
    matches = []
    only_in_times = []
    only_in_attr = []

    # Compare attributes present in both
    all_attrs = set(times_info.keys()) | set(attr_master.keys())

    for attr in sorted(all_attrs):
        ti = times_info.get(attr)
        am = attr_master.get(attr)

        if ti is None:
            only_in_attr.append(attr)
            continue
        if am is None:
            only_in_times.append(attr)
            continue

        ti_other = set(normalize_index_name(x) for x in ti["other_indexes"])
        am_other = set(am["other_indexes"])

        if ti_other != am_other:
            discrepancies.append({
                "attribute": attr,
                "times_info": ti["other_indexes"],
                "attr_master": am["other_indexes"],
                "ti_normalized": ti_other,
                "am_normalized": am_other,
            })
        elif ti_other:  # Both match and non-empty
            matches.append({
                "attribute": attr,
                "other_indexes": list(ti_other),
            })

    # Print discrepancies
    print(f"DISCREPANCIES: {len(discrepancies)}")
    print("-" * 80)
    for d in discrepancies:
        print(f"\n{d['attribute']}:")
        print(f"  times-info.json:           {d['times_info']} -> normalized: {d['ti_normalized']}")
        print(f"  Attribute-master-export:   {d['attr_master']} -> normalized: {d['am_normalized']}")

        # Show what's different
        only_ti = d['ti_normalized'] - d['am_normalized']
        only_am = d['am_normalized'] - d['ti_normalized']
        if only_ti:
            print(f"  Only in times-info:        {only_ti}")
        if only_am:
            print(f"  Only in Attribute-master:  {only_am}")

    print()
    print(f"\nMATCHES (both agree on non-empty other_indexes): {len(matches)}")
    print("-" * 80)
    for m in matches[:10]:  # Show first 10
        print(f"  {m['attribute']}: {m['other_indexes']}")
    if len(matches) > 10:
        print(f"  ... and {len(matches) - 10} more")

    print()
    print(f"\nATTRIBUTES ONLY IN times-info.json: {len(only_in_times)}")
    if only_in_times:
        print(f"  {', '.join(only_in_times[:20])}")
        if len(only_in_times) > 20:
            print(f"  ... and {len(only_in_times) - 20} more")

    print()
    print(f"\nATTRIBUTES ONLY IN Attribute-master-export.xlsx: {len(only_in_attr)}")
    if only_in_attr:
        print(f"  {', '.join(only_in_attr[:20])}")
        if len(only_in_attr) > 20:
            print(f"  ... and {len(only_in_attr) - 20} more")

    return discrepancies


if __name__ == "__main__":
    compare_other_indexes()
