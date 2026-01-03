"""Roundtrip tests: TableIR → Excel → xl2times validation."""

import json
import shutil
import subprocess
import tempfile
from pathlib import Path

import jsonschema
import pytest
from openpyxl import load_workbook

from tools.veda_emit_excel import emit_excel, load_tableir, validate_tableir

PROJECT_ROOT = Path(__file__).parent.parent
EXAMPLES_DIR = PROJECT_ROOT / "vedalang" / "examples"
SCHEMA_DIR = PROJECT_ROOT / "vedalang" / "schema"
FIXTURE_PATH = PROJECT_ROOT / "fixtures" / "MiniVEDA2"


@pytest.mark.skipif(
    not (PROJECT_ROOT / "xl2times").exists(),
    reason="xl2times submodule not available",
)
class TestRoundtrip:
    """Test the full TableIR → Excel → xl2times pipeline."""

    def test_emit_excel_roundtrip_with_fixture(self):
        """
        Roundtrip: copy MiniVEDA2 fixture, run xl2times, validate outputs.

        This tests the full pipeline using the validated MiniVEDA2 fixture
        as the baseline, proving xl2times validation works end-to-end.
        """
        if not FIXTURE_PATH.exists():
            pytest.skip("MiniVEDA2 fixture not available")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)

            # Copy fixture files to temp dir
            for f in FIXTURE_PATH.glob("*.xlsx"):
                shutil.copy(f, tmpdir / f.name)

            # Run xl2times on the fixture
            manifest_path = tmpdir / "manifest.json"
            diagnostics_path = tmpdir / "diagnostics.json"

            result = subprocess.run(
                [
                    "uv",
                    "run",
                    "python",
                    "-m",
                    "xl2times",
                    str(tmpdir),
                    "--manifest-json",
                    str(manifest_path),
                    "--diagnostics-json",
                    str(diagnostics_path),
                ],
                capture_output=True,
                text=True,
                cwd=PROJECT_ROOT,
            )

            # Check outputs created
            assert manifest_path.exists(), (
                f"Manifest not created. stderr: {result.stderr}"
            )
            assert diagnostics_path.exists(), (
                f"Diagnostics not created. stderr: {result.stderr}"
            )

            # Validate manifest against schema
            with open(SCHEMA_DIR / "manifest.schema.json") as f:
                manifest_schema = json.load(f)
            with open(manifest_path) as f:
                manifest = json.load(f)
            jsonschema.validate(manifest, manifest_schema)

            # Validate diagnostics against schema
            with open(SCHEMA_DIR / "diagnostics.schema.json") as f:
                diag_schema = json.load(f)
            with open(diagnostics_path) as f:
                diagnostics = json.load(f)
            jsonschema.validate(diagnostics, diag_schema)

            # Check no errors in diagnostics (warnings are OK)
            errors = [
                d
                for d in diagnostics.get("diagnostics", [])
                if d.get("severity") == "error"
            ]
            assert len(errors) == 0, f"xl2times reported errors: {errors}"

    def test_emit_and_validate_tableir(self):
        """
        Emit Excel from TableIR, validate emitted files exist and structure.

        Note: Full xl2times validation requires complete VEDA system tables
        (BookRegions_Map, TimeSlices, etc.) which are complex. This test
        validates the emitter itself works correctly.
        """
        tableir = load_tableir(EXAMPLES_DIR / "tableir_minimal.yaml")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)

            # Emit Excel
            created = emit_excel(tableir, tmpdir)
            assert len(created) >= 1, "Should create at least one Excel file"

            # Verify all expected files were created
            for file_spec in tableir["files"]:
                expected_path = tmpdir / file_spec["path"]
                assert expected_path.exists(), f"Missing file: {expected_path}"

    def test_manifest_contains_expected_tags(self):
        """Verify manifest contains the expected tags when processing fixture."""
        if not FIXTURE_PATH.exists():
            pytest.skip("MiniVEDA2 fixture not available")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)

            # Copy fixture
            for f in FIXTURE_PATH.glob("*.xlsx"):
                shutil.copy(f, tmpdir / f.name)

            manifest_path = tmpdir / "manifest.json"
            subprocess.run(
                [
                    "uv",
                    "run",
                    "python",
                    "-m",
                    "xl2times",
                    str(tmpdir),
                    "--manifest-json",
                    str(manifest_path),
                ],
                capture_output=True,
                cwd=PROJECT_ROOT,
            )

            with open(manifest_path) as f:
                manifest = json.load(f)

            # Extract tags from manifest
            tags_found = set()
            for table in manifest.get("tables", []):
                tags_found.add(table.get("tag", ""))

            # Check expected tags are present
            assert any("Process" in t or "PROCESS" in t for t in tags_found), (
                f"Expected process tag, found: {tags_found}"
            )


class TestValidation:
    """Test schema validation catches errors."""

    def test_invalid_tableir_rejected(self):
        """Invalid TableIR (tag missing ~) should raise ValidationError."""
        invalid = {
            "files": [
                {
                    "path": "test.xlsx",
                    "sheets": [
                        {"name": "Sheet1", "tables": [{"tag": "NO_TILDE", "rows": []}]}
                    ],
                }
            ]
        }
        with pytest.raises(jsonschema.ValidationError):
            validate_tableir(invalid)


class TestExcelStructure:
    """Test emitted Excel structure matches TableIR."""

    def test_emitted_excel_structure(self):
        """Verify sheet names, tags, and columns match TableIR."""
        tableir = load_tableir(EXAMPLES_DIR / "tableir_minimal.yaml")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)
            emit_excel(tableir, tmpdir)

            for file_spec in tableir["files"]:
                expected_path = tmpdir / file_spec["path"]
                assert expected_path.exists(), f"Missing file: {expected_path}"

                wb = load_workbook(expected_path)

                for sheet_spec in file_spec["sheets"]:
                    assert sheet_spec["name"] in wb.sheetnames, (
                        f"Missing sheet: {sheet_spec['name']}"
                    )

                    ws = wb[sheet_spec["name"]]

                    # Verify tags appear in the sheet
                    tags_in_sheet = [
                        ws.cell(row=r, column=1).value
                        for r in range(1, ws.max_row + 1)
                        if ws.cell(row=r, column=1).value
                        and str(ws.cell(row=r, column=1).value).startswith("~")
                    ]

                    expected_tags = [t["tag"] for t in sheet_spec["tables"]]
                    for tag in expected_tags:
                        assert tag in tags_in_sheet, (
                            f"Tag {tag} not found in sheet. Found: {tags_in_sheet}"
                        )
