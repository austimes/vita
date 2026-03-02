import tempfile
from pathlib import Path

import jsonschema
import pytest
from openpyxl import load_workbook

from tools.veda_emit_excel import emit_excel, load_tableir, validate_tableir

PROJECT_ROOT = Path(__file__).parent.parent
EXAMPLES_DIR = PROJECT_ROOT / "vedalang" / "examples"


def test_emit_minimal_tableir():
    """Emit tableir/tableir_minimal.yaml and verify Excel structure."""
    tableir = load_tableir(EXAMPLES_DIR / "tableir/tableir_minimal.yaml")

    with tempfile.TemporaryDirectory() as tmpdir:
        created = emit_excel(tableir, Path(tmpdir))

        assert len(created) >= 1
        for path in created:
            assert path.exists()
            wb = load_workbook(path)
            assert len(wb.sheetnames) > 0


def test_excel_contains_tag():
    """Verify emitted Excel contains the table tag."""
    tableir = {
        "files": [
            {
                "path": "test.xlsx",
                "sheets": [
                    {
                        "name": "Sheet1",
                        "tables": [{"tag": "~FI_TEST", "rows": [{"col1": "value1"}]}],
                    }
                ],
            }
        ]
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        created = emit_excel(tableir, Path(tmpdir))
        wb = load_workbook(created[0])
        ws = wb.active
        assert ws.cell(1, 1).value == "~FI_TEST"
        assert ws.cell(2, 1).value == "col1"
        assert ws.cell(3, 1).value == "value1"


def test_invalid_tableir_rejected():
    """Invalid TableIR should raise ValidationError."""
    invalid = {
        "files": [
            {
                "path": "x.xlsx",
                "sheets": [
                    {"name": "S", "tables": [{"tag": "NO_TILDE", "rows": []}]}
                ],
            }
        ]
    }
    with pytest.raises(jsonschema.ValidationError):
        validate_tableir(invalid)


def test_uc_sets_emitted_before_table():
    """Tables with uc_sets should emit ~UC_SETS declarations before the table tag."""
    tableir = {
        "files": [
            {
                "path": "uc_test.xlsx",
                "sheets": [
                    {
                        "name": "Constraints",
                        "tables": [
                            {
                                "tag": "~UC_T",
                                "uc_sets": {"R_E": "AllRegions", "T_E": ""},
                                "rows": [{"uc_n": "TEST", "value": 1}],
                            }
                        ],
                    }
                ],
            }
        ]
    }

    with tempfile.TemporaryDirectory() as tmpdir:
        created = emit_excel(tableir, Path(tmpdir))
        wb = load_workbook(created[0])
        ws = wb.active

        # Row 1: ~UC_SETS: R_E: AllRegions
        assert ws.cell(1, 1).value == "~UC_SETS: R_E: AllRegions"
        # Row 2: ~UC_SETS: T_E (no trailing colon/space for empty value)
        assert ws.cell(2, 1).value == "~UC_SETS: T_E"
        # Row 3: ~UC_T
        assert ws.cell(3, 1).value == "~UC_T"
        # Row 4: header
        assert ws.cell(4, 1).value == "uc_n"
